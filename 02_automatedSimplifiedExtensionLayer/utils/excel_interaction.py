from pathlib import Path
from openpyxl import load_workbook
from typing import Union

def get_files_in_directory(directory: Path) -> None:
    """
    Function to get files from a specified directory.

    Parameters
    ----------
    directory: str
        full path to directory

    Returns
    -------
    files: list
        a list with the files in directory

    """

    path = directory
    files = [Path(directory).glob("*")]
    return files


def read_from_excel_cell(path: Path, sheet: str|int, row: int, column: int) -> str:
    """
    Read a value in a certain cell of an excel file at a location given by path.

    Parameters
    ----------
    path: str
        location of excel file
    sheet: str
        sheet of excel file
    row: int
        row of excel sheet
    column: int
        column of excel sheet 
    Returns
    ----------
    str
        value to be read
    """
    # Convert into Path if it is no Path 
    if isinstance(path, Path):
        pass
    else:
        path = Path(path)

    workbook = load_workbook(workbook) # or load(workbook)

    # Use match case to get the specified sheet
    match sheet:
        case int():
            sheet = workbook[sheet_name - 1]
        case str():
            sheet = workbook[sheet_name]
        case _:
            raise ValueError("Invalid sheet name type. Should be int or str.")
    excel_cell_dict = {key:value for key, value in zip(range(1,28), list('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))}
    cell_address = f"{excel_cell_dict[column]}{row}"
    return sheet[cell_address]

def write_to_excel_cell(file_path: str|Path, sheet_name: str | int, cell_address: str, new_value: str | float | int) -> None:
    """
    Function that loads the ``sheet_name`` from the file lying at the location ``file_path`` to overwrite everything in
    ``cell_address`` with ``new_value``. (This function cannot be used to write content to a not yet existing sheet.)

    Parameters
    ----------
    file_path: str | pathlib.Path
    sheet_name: str|int
    cell_addresss: str
        one excel cell, e.g. 'A1'
    new_value: str|float|int
    """

    # Convert file path to Path object
    file_path = Path(file_path)
    
    # Load the workbook
    workbook = load_workbook(file_path)
    
    # Use match case to get the specified sheet
    match sheet_name:
        case int():
            sheet = workbook[sheet_name - 1]
        case str():
            sheet = workbook[sheet_name]
        case _:
            raise ValueError("Invalid sheet name type. Should be int or str.")
    
    # Update the cell value
    sheet[cell_address] = new_value
    
    # Save the changes
    workbook.save(file_path)

if __name__ == "__main__":
    pass
